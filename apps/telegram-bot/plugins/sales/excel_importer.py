from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from plugins.sales.models import PriceRow

REQUIRED = {"route", "min_kg", "sell_rate"}
OPTIONAL_ALIASES = {
    "route": ("route", "tuyen", "tuyến"),
    "carrier": ("carrier", "hang", "hãng"),
    "cargo": ("cargo", "cargo_type", "loai_hang", "loại hàng"),
    "min_kg": ("min_kg", "min", "weight_min_kg", "từ_kg", "tu_kg"),
    "max_kg": ("max_kg", "max", "weight_max_kg", "đến_kg", "den_kg"),
    "buy_rate": ("buy_rate", "gia_mua", "buy"),
    "sell_rate": ("sell_rate", "gia_ban", "sell", "price_per_kg", "gia"),
    "currency": ("currency", "tien_te", "tiền tệ"),
    "fsc_pct": ("fsc_%", "fsc_pct", "fsc"),
    "awb_fee": ("awb_fee", "phi_awb", "phí awb"),
    "effective_from": ("effective_from", "tu_ngay", "từ ngày"),
}


class ExcelImportError(ValueError):
    pass


def _norm_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("%", "_pct")
    text = re.sub(r"[\s\-]+", "_", text)
    return text


def _build_header_map(headers: list[Any]) -> dict[str, int]:
    normalized = [_norm_header(h) for h in headers]
    mapping: dict[str, int] = {}
    for field, aliases in OPTIONAL_ALIASES.items():
        for alias in aliases:
            key = _norm_header(alias)
            if key in normalized:
                mapping[field] = normalized.index(key)
                break
    return mapping


def _cell(row: tuple[Any, ...], idx: int | None) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _num(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    try:
        return float(text)
    except ValueError:
        return default


def _str(value: Any, default: str = "") -> str:
    if value is None:
        return default
    return str(value).strip()


def parse_tariff_excel(data: bytes) -> list[PriceRow]:
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ExcelImportError("File Excel không có sheet.")

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ExcelImportError("File Excel trống.")

    header_map = _build_header_map(list(header_row))
    missing = REQUIRED - set(header_map)
    if missing:
        raise ExcelImportError(
            "Thiếu cột bắt buộc: "
            + ", ".join(sorted(missing))
            + ". Cần: Route, Min_kg, Sell_rate."
        )

    out: list[PriceRow] = []
    for line_no, row in enumerate(rows_iter, start=2):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        route = _str(_cell(row, header_map.get("route"))).upper()
        if not route:
            continue

        sell = _num(_cell(row, header_map.get("sell_rate")))
        buy = _num(_cell(row, header_map.get("buy_rate")))
        price = sell if sell > 0 else buy
        if price <= 0:
            raise ExcelImportError(f"Dòng {line_no}: Sell_rate/Buy_rate phải > 0.")

        min_kg = _num(_cell(row, header_map.get("min_kg")))
        max_kg = _num(_cell(row, header_map.get("max_kg")))
        cargo = _str(_cell(row, header_map.get("cargo")), "general").lower() or "general"
        currency = _str(_cell(row, header_map.get("currency")), "VND") or "VND"
        carrier = _str(_cell(row, header_map.get("carrier")))
        fsc = _num(_cell(row, header_map.get("fsc_pct")))
        awb = _num(_cell(row, header_map.get("awb_fee")))
        effective = _str(_cell(row, header_map.get("effective_from")))

        notes_parts = []
        if carrier:
            notes_parts.append(f"carrier={carrier}")
        if fsc:
            notes_parts.append(f"FSC={fsc}%")
        if awb:
            notes_parts.append(f"AWB={awb:,.0f}")
        if effective:
            notes_parts.append(f"from={effective}")

        weight_range = ""
        if max_kg > 0:
            weight_range = f"{min_kg:g}-{max_kg:g}kg"
        elif min_kg > 0:
            weight_range = f"{min_kg:g}+kg"

        out.append(
            PriceRow(
                route=route,
                cargo_type=cargo,
                weight_min_kg=min_kg,
                weight_max_kg=max_kg,
                weight_range=weight_range,
                price_per_kg=price,
                currency=currency,
                notes="; ".join(notes_parts),
            )
        )

    wb.close()
    if not out:
        raise ExcelImportError("Không có dòng giá hợp lệ trong file.")
    return out
