from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

# Giới hạn để tiết kiệm token
_MAX_SHEETS = 3
_MAX_ROWS = 40
_MAX_COLS = 12
_MAX_CELL_CHARS = 80
_MAX_OUT_CHARS = 3500


def _cell_str(value) -> str:
    if value is None:
        return ""
    s = str(value).strip().replace("\n", " ")
    if len(s) > _MAX_CELL_CHARS:
        return s[: _MAX_CELL_CHARS - 1] + "…"
    return s


def excel_to_text(data: bytes, filename: str = "file.xlsx") -> str:
    """Đọc .xlsx → text bảng TSV ngắn (không gửi binary lên LLM)."""
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    parts: list[str] = [f"File: {filename}"]
    all_sheets = list(wb.sheetnames)
    sheets = all_sheets[:_MAX_SHEETS]
    for name in sheets:
        ws = wb[name]
        parts.append(f"\n## Sheet: {name}")
        row_n = 0
        for row in ws.iter_rows(max_row=_MAX_ROWS, max_col=_MAX_COLS, values_only=True):
            cells = [_cell_str(c) for c in row]
            if not any(cells):
                continue
            parts.append("\t".join(cells))
            row_n += 1
            if row_n >= _MAX_ROWS:
                parts.append("… (cắt bớt dòng)")
                break
        if not row_n:
            parts.append("(sheet trống)")
    wb.close()

    if len(all_sheets) > _MAX_SHEETS:
        parts.append(f"\n… còn {len(all_sheets) - _MAX_SHEETS} sheet khác (đã bỏ)")

    text = "\n".join(parts).strip()
    if len(text) > _MAX_OUT_CHARS:
        return text[: _MAX_OUT_CHARS - 1] + "…"
    return text


def is_excel_filename(name: str | None) -> bool:
    if not name:
        return False
    low = name.lower()
    return low.endswith(".xlsx") or low.endswith(".xlsm")


def is_excel_mime(mime: str | None) -> bool:
    if not mime:
        return False
    m = mime.lower()
    return m in {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel.sheet.macroenabled.12",
        "application/vnd.ms-excel",
    }
